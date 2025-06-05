from PyQt5.QtWidgets import (QWidget, QLabel, QVBoxLayout, QPushButton, QDialog, QLineEdit, QComboBox,
                             QFileDialog, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem,
                             QTableWidget, QTableWidgetItem, QAbstractItemView, QHeaderView,
                             QCheckBox, QMessageBox, QSpinBox, QFormLayout, QHBoxLayout,
                             QDoubleSpinBox,  # Added QDoubleSpinBox
                             QGroupBox, QSplitter, QScrollArea, QTextEdit, QListWidget,
                             QToolButton, QSizePolicy, QStyledItemDelegate, QTabWidget)  # Added QTabWidget
from PyQt5.QtGui import QPixmap, QImage, QPainter, QPen, QColor, QBrush, QCursor, QFont, QIcon, QPalette
from PyQt5.QtCore import Qt, QRectF, QPointF, pyqtSignal, QSizeF, QSize

import fitz
import os
from core_logic import (PDFProcessor, TemplateDefinition, FieldExtractionDefinition,
                        PYMUPDF_SUPPORTS_REDACT_ANNOT)

# --- Icons ---
ICON_PREV = "SP_ArrowLeft"
ICON_NEXT = "SP_ArrowRight"
ICON_UPLOAD = "SP_DialogOpenButton"
ICON_ADD = "SP_FileDialogNewFolder"
ICON_EDIT = "SP_FileDialogDetailedView"
ICON_DELETE = "SP_TrashIcon"

# --- Colors ---
DARK_BLUE_GRAY = "#435967"
LIGHT_GRAY_BLUE = "#E4E8EA"
HSBC_RED = "#BA1110"
TEAL = "#00847F"
MEDIUM_BLUE = "#305A86"
HIGHLIGHT_ORANGE = QColor(255, 165, 0)
ANCHOR_HIGHLIGHT_COLOR = QColor(HSBC_RED)


class PDFViewerWidget(QGraphicsView):  # No changes from previous fully correct version
    pointSelected = pyqtSignal(QPointF)
    boxSelected = pyqtSignal(QRectF)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setScene(QGraphicsScene(self))
        self.setRenderHint(QPainter.Antialiasing);
        self.setRenderHint(QPainter.SmoothPixmapTransform)
        self.setDragMode(QGraphicsView.ScrollHandDrag);
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorViewCenter);
        self._pixmap_item = None
        self._current_pdf_page_rect = None;
        self._render_zoom_factor = 2.0
        self._selection_mode = None;
        self._start_pos_view = None
        self._current_selection_rect_item = None;
        self.current_pdf_path = None
        self.current_page_idx = 0;
        self.pdf_processor = None
        self.current_redaction_texts = [];
        self.highlight_items = []
        self.setFocusPolicy(Qt.StrongFocus)

    def load_pdf(self, pdf_path, page_idx=0, redaction_texts=None):
        self.current_pdf_path = pdf_path;
        self.current_page_idx = page_idx
        self.current_redaction_texts = redaction_texts if redaction_texts is not None else []
        if self.pdf_processor:
            self.pdf_processor.load_pdf(pdf_path)
        elif pdf_path and os.path.exists(pdf_path):
            self.pdf_processor = PDFProcessor(pdf_path)
        else:
            self.pdf_processor = None
        if not self.pdf_processor or not self.pdf_processor.doc:
            self.scene().clear();
            self._pixmap_item = None
            if pdf_path: QMessageBox.warning(self, "PDF Error", f"Could not open or process PDF: {pdf_path}")
            return
        self.display_page(page_idx)

    def display_page(self, page_idx):
        if not self.pdf_processor or not self.pdf_processor.doc: self.scene().clear(); self._pixmap_item = None; return
        if not (
                0 <= page_idx < self.pdf_processor.get_page_count()): self.scene().clear(); self._pixmap_item = None; return
        self.current_page_idx = page_idx
        pix_or_qimage = self.pdf_processor.get_page_pixmap(page_idx, zoom=self._render_zoom_factor,
                                                           redaction_texts=self.current_redaction_texts)
        if pix_or_qimage:
            qpixmap = QPixmap.fromImage(pix_or_qimage) if isinstance(pix_or_qimage, QImage) else QPixmap.fromImage(
                QImage(pix_or_qimage.samples, pix_or_qimage.width, pix_or_qimage.height, pix_or_qimage.stride,
                       QImage.Format_RGB888 if not pix_or_qimage.alpha else QImage.Format_RGBA8888))
            current_transform = self.transform()
            self.scene().clear();
            self.highlight_items.clear()
            if self._current_selection_rect_item: self._current_selection_rect_item = None
            self._pixmap_item = self.scene().addPixmap(qpixmap)
            new_scene_rect = QRectF(qpixmap.rect())
            self.setSceneRect(new_scene_rect)
            page_obj = self.pdf_processor.get_page_object(page_idx)
            self._current_pdf_page_rect = page_obj.rect if page_obj else None
            if (
                    current_transform.m11() == 1.0 and current_transform.m22() == 1.0 and not self.sceneRect().isEmpty()) or self.sceneRect().isEmpty():
                self.fitInView(new_scene_rect, Qt.KeepAspectRatio)
            else:
                self.setTransform(current_transform)
        else:
            self.scene().clear(); self._pixmap_item = None

    def update_redactions(self, redaction_texts):
        self.current_redaction_texts = redaction_texts
        if self.pdf_processor and self.pdf_processor.doc: self.display_page(self.current_page_idx)

    def fit_to_view(self):
        if not self._pixmap_item or self.sceneRect().isEmpty(): return
        self.fitInView(self.sceneRect(), Qt.KeepAspectRatio)

    def set_selection_mode(self, mode):
        self._selection_mode = mode
        if mode:
            self.setDragMode(QGraphicsView.NoDrag); self.setCursor(Qt.CrossCursor)
        else:
            self.setDragMode(QGraphicsView.ScrollHandDrag);
            self.setCursor(Qt.ArrowCursor)
            if self._current_selection_rect_item and self._current_selection_rect_item.scene() == self.scene():
                self.scene().removeItem(self._current_selection_rect_item)
            self._current_selection_rect_item = None

    def mousePressEvent(self, event):
        if self._selection_mode and self._pixmap_item and event.button() == Qt.LeftButton:
            self._start_pos_view = event.pos();
            scene_start_pos = self.mapToScene(self._start_pos_view)
            if not self.sceneRect().contains(scene_start_pos): super().mousePressEvent(event); return
            if self._selection_mode == "point":
                pdf_point = self.scene_to_pdf_coords(scene_start_pos)
                if pdf_point: self.pointSelected.emit(pdf_point)
            elif self._selection_mode == "box":
                if self._current_selection_rect_item and self._current_selection_rect_item.scene() == self.scene():
                    self.scene().removeItem(self._current_selection_rect_item)
                self._current_selection_rect_item = self.scene().addRect(QRectF(scene_start_pos, QSizeF(0, 0)),
                                                                         QPen(QColor(HSBC_RED), 2, Qt.SolidLine));
                self._current_selection_rect_item.setZValue(10)
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._selection_mode == "box" and self._start_pos_view and self._current_selection_rect_item:
            self._current_selection_rect_item.setRect(
                QRectF(self.mapToScene(self._start_pos_view), self.mapToScene(event.pos())).normalized())
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._selection_mode == "box" and self._start_pos_view and self._current_selection_rect_item and event.button() == Qt.LeftButton:
            scene_rect = self._current_selection_rect_item.rect()
            if scene_rect.width() > 1 and scene_rect.height() > 1:
                pdf_rect = self.scene_to_pdf_coords_rect(scene_rect)
                if pdf_rect: self.boxSelected.emit(pdf_rect)
        self._start_pos_view = None;
        super().mouseReleaseEvent(event)

    def scene_to_pdf_coords(self, scene_pos: QPointF):
        if not self._pixmap_item or not self._current_pdf_page_rect or self.sceneRect().isEmpty() or not self.sceneRect().contains(
            scene_pos): return None
        rel_x = (scene_pos.x() - self.sceneRect().left()) / self.sceneRect().width();
        rel_y = (scene_pos.y() - self.sceneRect().top()) / self.sceneRect().height()
        return QPointF(self._current_pdf_page_rect.x0 + rel_x * self._current_pdf_page_rect.width,
                       self._current_pdf_page_rect.y0 + rel_y * self._current_pdf_page_rect.height)

    def scene_to_pdf_coords_rect(self, scene_rect: QRectF):
        tl_pdf = self.scene_to_pdf_coords(scene_rect.topLeft());
        br_pdf = self.scene_to_pdf_coords(scene_rect.bottomRight())
        if tl_pdf and br_pdf: return QRectF(tl_pdf, br_pdf).normalized()
        return None

    def pdf_to_scene_coords_rect(self, pdf_coords_tuple):
        if not self._pixmap_item or not self._current_pdf_page_rect or self.sceneRect().isEmpty() or self._current_pdf_page_rect.width == 0 or self._current_pdf_page_rect.height == 0: return None
        pdf_rect = fitz.Rect(pdf_coords_tuple);
        pr = self._current_pdf_page_rect;
        sr_base = self.sceneRect()
        rel_x0 = (pdf_rect.x0 - pr.x0) / pr.width;
        rel_y0 = (pdf_rect.y0 - pr.y0) / pr.height;
        rel_x1 = (pdf_rect.x1 - pr.x0) / pr.width;
        rel_y1 = (pdf_rect.y1 - pr.y0) / pr.height
        return QRectF(QPointF(sr_base.left() + rel_x0 * sr_base.width(), sr_base.top() + rel_y0 * sr_base.height()),
                      QPointF(sr_base.left() + rel_x1 * sr_base.width(),
                              sr_base.top() + rel_y1 * sr_base.height())).normalized()

    def add_highlight_multi(self, value_coords_list, anchor_coords_list=None, value_color=HIGHLIGHT_ORANGE,
                            anchor_color=ANCHOR_HIGHLIGHT_COLOR, opacity=0.35):
        if not self._pixmap_item: return

        def draw_rects(coords_list, color, z_val):
            if not coords_list: return
            if not isinstance(coords_list, list): coords_list = [coords_list]
            for pdf_coords_tuple in coords_list:
                if not pdf_coords_tuple: continue
                scene_rect = self.pdf_to_scene_coords_rect(pdf_coords_tuple)
                if scene_rect:
                    if self._current_selection_rect_item and z_val == 1 and abs(
                            self._current_selection_rect_item.rect().left() - scene_rect.left()) < 0.01 and abs(
                            self._current_selection_rect_item.rect().width() - scene_rect.width()) < 0.01:
                        if self._current_selection_rect_item.scene() == self.scene(): self.scene().removeItem(
                            self._current_selection_rect_item)
                        self._current_selection_rect_item = None
                    brush_color = QColor(color);
                    brush_color.setAlphaF(opacity)
                    item = self.scene().addRect(scene_rect, QPen(Qt.NoPen), QBrush(brush_color));
                    item.setZValue(z_val);
                    self.highlight_items.append(item)

        draw_rects(anchor_coords_list, anchor_color, 0)
        draw_rects(value_coords_list, value_color, 1)

    def clear_highlights(self):
        for item in self.highlight_items:
            if item.scene() == self.scene(): self.scene().removeItem(item)
        self.highlight_items.clear()
        if self._current_selection_rect_item and self._current_selection_rect_item.scene() == self.scene():
            self.scene().removeItem(self._current_selection_rect_item)
        self._current_selection_rect_item = None

    def wheelEvent(self, event):
        if event.modifiers() == Qt.ControlModifier:
            factor = 1.15 if event.angleDelta().y() > 0 else 1 / 1.15
            self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse);
            self.scale(factor, factor);
            event.accept()
        else:
            super().wheelEvent(event)

    def keyPressEvent(self, event):
        if event.modifiers() == Qt.ControlModifier:
            if event.key() == Qt.Key_Plus or event.key() == Qt.Key_Equal:
                self.scale(1.2, 1.2)
            elif event.key() == Qt.Key_Minus:
                self.scale(1 / 1.2, 1 / 1.2)
            elif event.key() == Qt.Key_0:
                self.fit_to_view()
            else:
                super().keyPressEvent(event)
        else:
            super().keyPressEvent(event)


class FieldExtractionDialog(QDialog):  # Modified to not have its own PDF viewer
    fieldDefinitionSaved = pyqtSignal(FieldExtractionDefinition)

    def __init__(self, parent_template: TemplateDefinition,
                 field_definition: FieldExtractionDefinition = None,
                 # No pdf_processor or current_page_idx needed from parent anymore
                 # It will use the main viewer's state via TemplateConfigurationTab
                 parent_config_tab,  # Pass the TemplateConfigurationTab instance
                 parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Define Field for '{parent_template.name}'")
        self.setMinimumSize(550, 650)  # Smaller, as no PDF viewer
        self.setStyleSheet(f"QDialog {{ background-color: {LIGHT_GRAY_BLUE}; }}")

        self.parent_template = parent_template
        self.current_field_def = field_definition if field_definition else FieldExtractionDefinition()
        self.parent_config_tab = parent_config_tab  # Reference to TemplateConfigurationTab

        # Main layout is just the controls now
        self.controls_layout_main = QVBoxLayout(self)  # Renamed to avoid conflict

        self.controls_scroll_area = QScrollArea()
        self.controls_scroll_area.setWidgetResizable(True)
        self.controls_widget = QWidget()
        self.dialog_controls_form_layout = QVBoxLayout(self.controls_widget)  # Renamed
        self.controls_widget.setLayout(self.dialog_controls_form_layout)
        self.controls_scroll_area.setWidget(self.controls_widget)

        self.controls_layout_main.addWidget(self.controls_scroll_area)

        self._setup_field_controls_in_dialog()  # Renamed method
        self._populate_from_field_def()

        # Connect signals for point/box selection from the *main* PDF viewer in TemplateConfigurationTab
        # These connections are now made/managed by TemplateConfigurationTab when dialog is shown

    def _setup_field_controls_in_dialog(self):  # Renamed
        form_layout = QFormLayout()
        form_layout.setContentsMargins(10, 10, 10, 10);
        form_layout.setSpacing(10)
        form_layout.setLabelAlignment(Qt.AlignRight)

        # Field ID (Read-only)
        self.field_id_label_dialog = QLabel(self.current_field_def.field_id)  # Suffix Dialog to avoid name clash
        form_layout.addRow("Field ID:", self.field_id_label_dialog)

        # Field Label
        self.field_label_edit_dialog = QLineEdit()
        self.field_label_edit_dialog.setPlaceholderText("e.g., Invoice Number, Total Amount")
        form_layout.addRow("Field Label:", self.field_label_edit_dialog)

        # Sub-Type
        self.sub_type_combo_dialog = QComboBox()
        self.sub_type_combo_dialog.addItems(["Anchor Word", "Anchor Point", "Coordinate Box"])
        self.sub_type_combo_dialog.currentTextChanged.connect(self.on_field_subtype_changed_dialog)
        form_layout.addRow("Extraction Type:", self.sub_type_combo_dialog)

        # --- Anchor Word Group ---
        self.anchor_word_group_dialog = QGroupBox("Anchor Word / Key Options")
        aw_layout = QFormLayout()
        self.aw_anchor_text_edit_dialog = QLineEdit()
        aw_layout.addRow("Anchor Text / Key:", self.aw_anchor_text_edit_dialog)
        self.aw_match_type_combo_dialog = QComboBox();
        self.aw_match_type_combo_dialog.addItems(["Exact Match", "Case Insensitive", "Regular Expression"])
        aw_layout.addRow("Match Type (for Anchor/Key):", self.aw_match_type_combo_dialog)
        self.aw_scope_combo_dialog = QComboBox();
        self.aw_scope_combo_dialog.addItems(
            ["Whole Line", "Next N Words (Right)", "Previous N Words (Left)", "Text above the Word",
             "Text below the Word", "Value After Anchor (Line)", "Value After Anchor (Next N Words)"])
        self.aw_scope_combo_dialog.currentTextChanged.connect(self._toggle_field_n_words_visibility_dialog)
        aw_layout.addRow("Extract:", self.aw_scope_combo_dialog)
        self.aw_n_words_spin_dialog = QSpinBox();
        self.aw_n_words_spin_dialog.setMinimum(0);
        self.aw_n_words_spin_dialog.setValue(3)  # Allow 0 for non-N scopes
        aw_layout.addRow("N Words (if applicable):", self.aw_n_words_spin_dialog)
        self.anchor_word_group_dialog.setLayout(aw_layout)

        # --- Anchor Point Group ---
        self.anchor_point_group_dialog = QGroupBox("Anchor Point Options")
        ap_layout = QFormLayout()
        self.ap_select_button_dialog = QPushButton("Select Point from Main Viewer")
        self.ap_select_button_dialog.clicked.connect(
            self.activate_point_selection_in_main_viewer)  # Signal to parent tab
        ap_layout.addRow(self.ap_select_button_dialog)
        self.ap_coords_label_dialog = QLabel("Not selected")
        ap_layout.addRow("Selected Point:", self.ap_coords_label_dialog)
        self.ap_scope_combo_dialog = QComboBox();
        self.ap_scope_combo_dialog.addItems(
            ["Whole Line", "Next N Words (Right)", "Previous N Words (Left)", "Text above the Point",
             "Text below the Point", "Text at/on the Point"])
        self.ap_scope_combo_dialog.currentTextChanged.connect(self._toggle_field_n_words_visibility_dialog)
        ap_layout.addRow("Extract:", self.ap_scope_combo_dialog)
        self.ap_n_words_spin_dialog = QSpinBox();
        self.ap_n_words_spin_dialog.setMinimum(0);
        self.ap_n_words_spin_dialog.setValue(3)  # Allow 0
        ap_layout.addRow("N Words (if applicable):", self.ap_n_words_spin_dialog)
        self.anchor_point_group_dialog.setLayout(ap_layout)

        # --- Coordinate Box Group ---
        self.coord_box_group_dialog = QGroupBox("Coordinate Box Options")
        cb_layout = QFormLayout()
        self.cb_draw_button_dialog = QPushButton("Draw Box in Main Viewer")
        self.cb_draw_button_dialog.clicked.connect(self.activate_box_selection_in_main_viewer)  # Signal to parent tab
        cb_layout.addRow(self.cb_draw_button_dialog)
        self.cb_coords_label_dialog = QLabel("Not drawn")
        cb_layout.addRow("Box Coords:", self.cb_coords_label_dialog)
        self.coord_box_group_dialog.setLayout(cb_layout)

        # --- Tolerances Group ---
        self.tolerance_group_dialog = QGroupBox("Text Above/Below Tolerances")
        tol_layout = QFormLayout()
        self.tol_y_spin_dialog = QSpinBox();
        self.tol_y_spin_dialog.setMinimum(5);
        self.tol_y_spin_dialog.setMaximum(200);
        self.tol_y_spin_dialog.setValue(30);
        self.tol_y_spin_dialog.setSuffix(" px/pt")
        tol_layout.addRow("Vertical Search (Y):", self.tol_y_spin_dialog)
        self.tol_x_factor_spin_dialog = QDoubleSpinBox();
        self.tol_x_factor_spin_dialog.setMinimum(1.0);
        self.tol_x_factor_spin_dialog.setMaximum(5.0);
        self.tol_x_factor_spin_dialog.setSingleStep(0.1);
        self.tol_x_factor_spin_dialog.setValue(1.5);
        self.tol_x_factor_spin_dialog.setSuffix(" x Anchor Width")
        tol_layout.addRow("Horizontal Factor (X):", self.tol_x_factor_spin_dialog)
        self.tolerance_group_dialog.setLayout(tol_layout)

        self.extracted_text_preview_dialog = QTextEdit();
        self.extracted_text_preview_dialog.setReadOnly(True);
        self.extracted_text_preview_dialog.setFixedHeight(80)

        self.dialog_controls_form_layout.addLayout(form_layout)  # Add the QFormLayout
        self.dialog_controls_form_layout.addWidget(self.anchor_word_group_dialog)
        self.dialog_controls_form_layout.addWidget(self.anchor_point_group_dialog)
        self.dialog_controls_form_layout.addWidget(self.coord_box_group_dialog)
        self.dialog_controls_form_layout.addWidget(self.tolerance_group_dialog)
        self.dialog_controls_form_layout.addWidget(QLabel("Preview Extracted Text:"))
        self.dialog_controls_form_layout.addWidget(self.extracted_text_preview_dialog)
        self.dialog_controls_form_layout.addStretch()

        button_layout_dialog = QHBoxLayout()
        self.test_field_button_dialog = QPushButton("Test This Field (on Main Viewer)");
        self.test_field_button_dialog.clicked.connect(self.test_current_field_on_main_viewer)
        self.save_field_button_dialog = QPushButton("OK / Save Field");
        self.save_field_button_dialog.clicked.connect(self.save_field);
        self.save_field_button_dialog.setDefault(True)
        self.cancel_button_dialog = QPushButton("Cancel");
        self.cancel_button_dialog.clicked.connect(self.reject)
        button_layout_dialog.addWidget(self.test_field_button_dialog);
        button_layout_dialog.addStretch()
        button_layout_dialog.addWidget(self.save_field_button_dialog);
        button_layout_dialog.addWidget(self.cancel_button_dialog)
        self.controls_layout_main.addLayout(button_layout_dialog)  # Add buttons to main QVBoxLayout of dialog

    def _populate_from_field_def(self):  # Suffix _dialog to widget names
        fd = self.current_field_def
        self.field_label_edit_dialog.setText(fd.label)
        self.sub_type_combo_dialog.setCurrentText(fd.sub_type or "Anchor Word")
        self.aw_anchor_text_edit_dialog.setText(fd.anchor_text or "")
        self.aw_match_type_combo_dialog.setCurrentText(fd.match_type or "Exact Match")
        self.aw_scope_combo_dialog.setCurrentText(fd.extraction_scope or "Whole Line")
        self.aw_n_words_spin_dialog.setValue(
            fd.n_words or (5 if "Value After Anchor (Next N Words)" in (fd.extraction_scope or "") else 3))
        if fd.anchor_point:
            self.ap_coords_label_dialog.setText(f"({fd.anchor_point[0]:.1f}, {fd.anchor_point[1]:.1f})")
        else:
            self.ap_coords_label_dialog.setText("Not selected")
        self.ap_scope_combo_dialog.setCurrentText(fd.extraction_scope or "Whole Line")
        self.ap_n_words_spin_dialog.setValue(fd.n_words or 3)
        if fd.coordinate_box:
            cb = fd.coordinate_box; self.cb_coords_label_dialog.setText(
                f"({cb[0]:.1f},{cb[1]:.1f}) to ({cb[2]:.1f},{cb[3]:.1f})")
        else:
            self.cb_coords_label_dialog.setText("Not drawn")
        self.tol_y_spin_dialog.setValue(fd.above_below_tolerance_y)
        self.tol_x_factor_spin_dialog.setValue(fd.above_below_tolerance_x_factor)
        self.on_field_subtype_changed_dialog(self.sub_type_combo_dialog.currentText())

    def _update_field_def_from_form(self):  # Suffix _dialog to widget names
        fd = self.current_field_def
        fd.label = self.field_label_edit_dialog.text()
        fd.sub_type = self.sub_type_combo_dialog.currentText()
        fd.anchor_text = None;
        fd.match_type = "Exact Match";
        fd.anchor_point = None;
        fd.coordinate_box = None  # Reset
        if fd.sub_type == "Anchor Word":
            fd.anchor_text = self.aw_anchor_text_edit_dialog.text()
            fd.match_type = self.aw_match_type_combo_dialog.currentText()
            fd.extraction_scope = self.aw_scope_combo_dialog.currentText()
            fd.n_words = self.aw_n_words_spin_dialog.value() if (
                        "N Words" in fd.extraction_scope or "Value After Anchor (Next N Words)" in fd.extraction_scope) else 0
        elif fd.sub_type == "Anchor Point":
            # anchor_point is set by parent_config_tab calling update_anchor_point_from_main_viewer
            fd.extraction_scope = self.ap_scope_combo_dialog.currentText()
            fd.n_words = self.ap_n_words_spin_dialog.value() if "N Words" in fd.extraction_scope else 0
        elif fd.sub_type == "Coordinate Box":
            # coordinate_box is set by parent_config_tab calling update_coord_box_from_main_viewer
            fd.extraction_scope = None;
            fd.n_words = 0
        fd.above_below_tolerance_y = self.tol_y_spin_dialog.value()
        fd.above_below_tolerance_x_factor = self.tol_x_factor_spin_dialog.value()
        return fd

    def on_field_subtype_changed_dialog(self, sub_type):  # Suffix _dialog
        self.anchor_word_group_dialog.setVisible(sub_type == "Anchor Word")
        self.anchor_point_group_dialog.setVisible(sub_type == "Anchor Point")
        self.coord_box_group_dialog.setVisible(sub_type == "Coordinate Box")
        self.parent_config_tab.template_pdf_viewer.clear_highlights()  # Clear main viewer highlights
        self.parent_config_tab.template_pdf_viewer.set_selection_mode(None)  # Reset main viewer mode
        self._toggle_field_n_words_visibility_dialog()

    def _toggle_field_n_words_visibility_dialog(self):  # Suffix _dialog
        current_subtype = self.sub_type_combo_dialog.currentText()
        scope = "";
        n_words_spin_widget = None;
        n_words_label_widget = None
        show_n_words = False;
        show_tolerance = False
        if current_subtype == "Anchor Word":
            scope = self.aw_scope_combo_dialog.currentText();
            n_words_spin_widget = self.aw_n_words_spin_dialog
            n_words_label_widget = self.anchor_word_group_dialog.layout().labelForField(n_words_spin_widget)
            show_n_words = "N Words" in scope or "Value After Anchor (Next N Words)" in scope
            show_tolerance = "above the Word" in scope or "below the Word" in scope
        elif current_subtype == "Anchor Point":
            scope = self.ap_scope_combo_dialog.currentText();
            n_words_spin_widget = self.ap_n_words_spin_dialog
            n_words_label_widget = self.anchor_point_group_dialog.layout().labelForField(n_words_spin_widget)
            show_n_words = "N Words" in scope
            show_tolerance = "above the Point" in scope or "below the Point" in scope
        if n_words_spin_widget:
            if n_words_label_widget: n_words_label_widget.setVisible(show_n_words)
            n_words_spin_widget.setVisible(show_n_words)
        self.tolerance_group_dialog.setVisible(show_tolerance)

    # Methods to ask parent (TemplateConfigurationTab) to set mode on its viewer
    def activate_point_selection_in_main_viewer(self):
        self.parent_config_tab.set_main_viewer_selection_mode("point", self)  # Pass self as listener

    def activate_box_selection_in_main_viewer(self):
        self.parent_config_tab.set_main_viewer_selection_mode("box", self)  # Pass self as listener

    # Callbacks for when selection is made in main viewer
    def update_anchor_point_from_main_viewer(self, pdf_point: QPointF):
        self.current_field_def.anchor_point = (pdf_point.x(), pdf_point.y())
        self.ap_coords_label_dialog.setText(f"({pdf_point.x():.1f}, {pdf_point.y():.1f})")
        # Highlight on main viewer is handled by parent_config_tab

    def update_coord_box_from_main_viewer(self, pdf_rect: QRectF):
        self.current_field_def.coordinate_box = (pdf_rect.left(), pdf_rect.top(), pdf_rect.right(), pdf_rect.bottom())
        cb = self.current_field_def.coordinate_box
        self.cb_coords_label_dialog.setText(f"({cb[0]:.1f},{cb[1]:.1f}) to ({cb[2]:.1f},{cb[3]:.1f})")

    def test_current_field_on_main_viewer(self):
        main_pdf_processor = self.parent_config_tab.pdf_processor
        current_main_page_idx = self.parent_config_tab.current_test_page_idx
        main_viewer = self.parent_config_tab.template_pdf_viewer

        if not main_pdf_processor or not main_pdf_processor.doc:
            QMessageBox.warning(self, "Test Field", "No Test PDF loaded in the main Template Configuration view.");
            return

        field_to_test = self._update_field_def_from_form()
        if not field_to_test.label:
            QMessageBox.warning(self, "Test Field", "Please provide a Field Label.");
            self.field_label_edit_dialog.setFocus();
            return

        try:
            result: ExtractedFieldResult = main_pdf_processor.extract_field_data(field_to_test, current_main_page_idx)
            self.extracted_text_preview_dialog.setText(result.extracted_text)
            main_viewer.clear_highlights()
            main_viewer.add_highlight_multi(result.highlight_rects_pdf_coords, result.anchor_highlight_rects_pdf_coords)
        except Exception as e:
            self.extracted_text_preview_dialog.setText(f"[ERROR]: {e}")
            QMessageBox.critical(self, "Test Error", f"Error during field test: {e}")

    def save_field(self):  # Same validation logic as before
        field_def = self._update_field_def_from_form()
        if not field_def.label: QMessageBox.warning(self, "Save Field",
                                                    "Field Label cannot be empty."); self.field_label_edit_dialog.setFocus(); return
        valid = True;
        msg = "";
        widget_to_focus = None
        if field_def.sub_type == "Anchor Word":
            if not field_def.anchor_text:
                valid = False; msg = "Anchor Text / Key is required."; widget_to_focus = self.aw_anchor_text_edit_dialog
            elif "Value After Anchor (Next N Words)" in field_def.extraction_scope and (field_def.n_words or 0) <= 0:
                valid = False;
                msg = "N Words must be > 0 for this scope.";
                widget_to_focus = self.aw_n_words_spin_dialog
        elif field_def.sub_type == "Anchor Point" and not field_def.anchor_point:
            valid = False; msg = "Anchor point must be selected."; widget_to_focus = self.ap_select_button_dialog
        elif field_def.sub_type == "Coordinate Box" and not field_def.coordinate_box:
            valid = False; msg = "Coordinate box must be drawn."; widget_to_focus = self.cb_draw_button_dialog
        if not valid: QMessageBox.warning(self, "Save Field Error", msg);
        if widget_to_focus: widget_to_focus.setFocus(); return

    self.fieldDefinitionSaved.emit(field_def)
    self.accept()


class TemplateConfigurationTab(QWidget):
    templates_updated = pyqtSignal(list)

    def __init__(self, templates_list_ref: list[TemplateDefinition], parent=None):
        super().__init__(parent)
        self.templates_list_ref = templates_list_ref
        self.current_template: TemplateDefinition = None
        self.pdf_processor = PDFProcessor()  # This is THE PDFProcessor for this tab
        self.current_test_page_idx = 0
        self.active_field_dialog: FieldExtractionDialog = None  # To route viewer signals
        self._init_ui_restructured()  # New UI init method
        self.load_and_display_templates_list()

    def _init_ui_restructured(self):
        main_splitter = QSplitter(Qt.Horizontal, self)
        main_layout = QHBoxLayout(self)
        main_layout.addWidget(main_splitter)

        # --- Left Panel: Template List ---
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.addWidget(QLabel("<b>Document Templates</b>"))
        template_list_buttons = QHBoxLayout()
        add_template_btn = QPushButton("New");
        add_template_btn.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_ADD, self.style().SP_FileDialogNewFolder))));
        add_template_btn.clicked.connect(self.add_new_template)
        self.delete_template_btn = QPushButton("Delete");
        self.delete_template_btn.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_DELETE, self.style().SP_TrashIcon))));
        self.delete_template_btn.clicked.connect(self.delete_selected_template);
        self.delete_template_btn.setEnabled(False)
        template_list_buttons.addWidget(add_template_btn);
        template_list_buttons.addWidget(self.delete_template_btn);
        template_list_buttons.addStretch()
        left_layout.addLayout(template_list_buttons)
        self.templates_list_widget = QListWidget();
        self.templates_list_widget.itemSelectionChanged.connect(self.on_template_selected)
        left_layout.addWidget(self.templates_list_widget)
        left_panel.setLayout(left_layout)
        main_splitter.addWidget(left_panel)

        # --- Center Panel: PDF Viewer ---
        pdf_viewer_panel = QWidget()
        pdf_viewer_layout = QVBoxLayout(pdf_viewer_panel)
        page_nav_layout = QHBoxLayout()
        self.prev_page_btn_viewer = QToolButton();
        self.prev_page_btn_viewer.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_PREV, self.style().SP_ArrowLeft))));
        self.prev_page_btn_viewer.clicked.connect(self.viewer_prev_page)
        self.page_indicator_label = QLabel("Page: -/- (Load Test PDF)")
        self.next_page_btn_viewer = QToolButton();
        self.next_page_btn_viewer.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_NEXT, self.style().SP_ArrowRight))));
        self.next_page_btn_viewer.clicked.connect(self.viewer_next_page)
        page_nav_layout.addWidget(self.prev_page_btn_viewer);
        page_nav_layout.addStretch();
        page_nav_layout.addWidget(self.page_indicator_label);
        page_nav_layout.addStretch();
        page_nav_layout.addWidget(self.next_page_btn_viewer)
        pdf_viewer_layout.addLayout(page_nav_layout)
        self.template_pdf_viewer = PDFViewerWidget()  # This is the main viewer
        self.template_pdf_viewer.pdf_processor = self.pdf_processor  # Share the processor
        self.template_pdf_viewer.pointSelected.connect(self.on_main_viewer_point_selected)  # Connect signals
        self.template_pdf_viewer.boxSelected.connect(self.on_main_viewer_box_selected)
        pdf_viewer_layout.addWidget(self.template_pdf_viewer)
        pdf_viewer_panel.setLayout(pdf_viewer_layout)
        main_splitter.addWidget(pdf_viewer_panel)

        # --- Right Panel: Tabbed Details ---
        right_tab_widget = QTabWidget()

        # Tab 1: Template Properties
        props_tab_widget = QWidget()
        props_layout = QVBoxLayout(props_tab_widget)
        template_props_group = QGroupBox("Template Properties")
        props_form = QFormLayout()
        self.template_id_label = QLabel("N/A");
        props_form.addRow("Template ID:", self.template_id_label)
        self.template_name_edit = QLineEdit();
        self.template_name_edit.editingFinished.connect(lambda: self.save_current_template_details(silent=True))
        props_form.addRow("Template Name:", self.template_name_edit)
        self.keywords_edit = QLineEdit();
        self.keywords_edit.setPlaceholderText("Comma-separated");
        self.keywords_edit.editingFinished.connect(lambda: self.save_current_template_details(silent=True))
        props_form.addRow("Identifying Keywords:", self.keywords_edit)
        test_pdf_layout = QHBoxLayout();
        self.test_pdf_label = QLabel("No Test PDF");
        test_pdf_layout.addWidget(self.test_pdf_label, 1)
        load_test_pdf_btn = QPushButton("Load PDF");
        load_test_pdf_btn.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_UPLOAD, self.style().SP_DialogOpenButton))))
        load_test_pdf_btn.clicked.connect(self.load_test_pdf_for_template);
        test_pdf_layout.addWidget(load_test_pdf_btn)
        props_form.addRow("Test PDF:", test_pdf_layout)
        self.redactions_edit = QLineEdit();
        self.redactions_edit.setPlaceholderText("Comma-separated");
        self.redactions_edit.editingFinished.connect(lambda: self.save_current_template_details(silent=True))
        props_form.addRow("Redaction Texts:", self.redactions_edit)
        save_props_btn = QPushButton("Save Template Properties");
        save_props_btn.clicked.connect(lambda: self.save_current_template_details(silent=False))  # Explicit save
        props_form.addRow(save_props_btn)
        template_props_group.setLayout(props_form)
        props_layout.addWidget(template_props_group);
        props_layout.addStretch()
        right_tab_widget.addTab(props_tab_widget, "Properties")

        # Tab 2: Field Definitions
        fields_tab_widget = QWidget()
        fields_layout = QVBoxLayout(fields_tab_widget)
        field_buttons_layout = QHBoxLayout()
        self.add_field_btn = QPushButton("Add Field");
        self.add_field_btn.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_ADD, self.style().SP_FileDialogNewFolder))));
        self.add_field_btn.clicked.connect(self.add_new_field_definition);
        self.add_field_btn.setEnabled(False)
        self.edit_field_btn = QPushButton("Edit Field");
        self.edit_field_btn.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_EDIT, self.style().SP_FileDialogDetailedView))));
        self.edit_field_btn.clicked.connect(self.edit_selected_field_definition);
        self.edit_field_btn.setEnabled(False)
        self.delete_field_btn = QPushButton("Delete Field");
        self.delete_field_btn.setIcon(
            QIcon(self.style().standardIcon(getattr(self.style(), ICON_DELETE, self.style().SP_TrashIcon))));
        self.delete_field_btn.clicked.connect(self.delete_selected_field_definition);
        self.delete_field_btn.setEnabled(False)
        field_buttons_layout.addWidget(self.add_field_btn);
        field_buttons_layout.addWidget(self.edit_field_btn);
        field_buttons_layout.addWidget(self.delete_field_btn);
        field_buttons_layout.addStretch()
        fields_layout.addLayout(field_buttons_layout)
        self.fields_table = QTableWidget();
        self.fields_table.setColumnCount(5);
        self.fields_table.setHorizontalHeaderLabels(["Field Label", "Type", "Anchor/Key", "Extraction Scope", "N"])
        self.fields_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.fields_table.setSelectionBehavior(QAbstractItemView.SelectRows);
        self.fields_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fields_table.itemSelectionChanged.connect(self.on_field_selection_changed)
        self.fields_table.doubleClicked.connect(
            lambda: self.edit_selected_field_definition() if self.edit_field_btn.isEnabled() else None)
        fields_layout.addWidget(self.fields_table)
        self.test_all_fields_btn = QPushButton("Test All Fields (on Current Page)");
        self.test_all_fields_btn.clicked.connect(self.test_all_fields_on_current_page);
        self.test_all_fields_btn.setEnabled(False)
        fields_layout.addWidget(self.test_all_fields_btn, 0, Qt.AlignRight)
        right_tab_widget.addTab(fields_tab_widget, "Field Definitions")

        main_splitter.addWidget(right_tab_widget)
        main_splitter.setSizes([200, 500, 300])  # Adjust sizes: List, PDF Viewer, Details Tabs
        self.set_template_details_enabled(False)

    # --- Methods from old _init_ui (load_and_display_templates_list, add_new_template, etc.) ---
    # These are largely the same, ensure they interact with new UI elements.
    # ... (All other methods from TemplateConfigurationTab in previous response, ensure they are correctly adapted to new UI elements and the single PDF viewer model) ...
    # Key changes needed in:
    # - add_new_field_definition / edit_selected_field_definition: Pass `self` to FieldExtractionDialog.
    # - Methods that interacted with the old PDF viewer group now use `self.template_pdf_viewer`.

    # --- Methods for TemplateConfigurationTab (add_new_template, on_template_selected, etc.) ---
    # Copied and adapted from previous full implementation of TemplateConfigurationTab.
    def set_template_details_enabled(self, enabled):
        # Enable/disable elements in the "Properties" tab
        self.template_name_edit.setEnabled(enabled)
        self.keywords_edit.setEnabled(enabled)
        self.redactions_edit.setEnabled(enabled)
        # The "Load PDF" button in properties should generally be enabled if a template is selected

        # Enable/disable elements in the "Field Definitions" tab
        self.add_field_btn.setEnabled(enabled)
        # edit/delete for fields table are handled by their own selection logic
        self.test_all_fields_btn.setEnabled(enabled and self.pdf_processor.doc is not None)

    def load_and_display_templates_list(self):
        self.templates_list_widget.clear()
        selected_template_id = self.current_template.template_id if self.current_template else None
        new_selected_row = -1
        for i, template in enumerate(self.templates_list_ref):
            item = QListWidgetItem(f"{template.name} (ID: {template.template_id})")
            item.setData(Qt.UserRole, template.template_id)
            self.templates_list_widget.addItem(item)
            if template.template_id == selected_template_id:
                new_selected_row = i
        self.delete_template_btn.setEnabled(
            self.templates_list_widget.count() > 0 and self.templates_list_widget.currentItem() is not None)
        if new_selected_row != -1:
            self.templates_list_widget.setCurrentRow(new_selected_row)
        elif self.templates_list_widget.count() > 0:
            self.templates_list_widget.setCurrentRow(0)
        else:
            self.clear_template_details_ui()

    def add_new_template(self):
        name, ok = QInputDialog.getText(self, "New Template", "Enter Unique Template Name:")
        if ok and name:
            if any(t.name.lower() == name.lower() for t in self.templates_list_ref):
                QMessageBox.warning(self, "Duplicate Name", "A template with this name already exists.");
                return
            new_template = TemplateDefinition(name=name)
            self.templates_list_ref.append(new_template)
            self.templates_updated.emit(self.templates_list_ref)
            self.load_and_display_templates_list()
            for i in range(self.templates_list_widget.count()):
                if self.templates_list_widget.item(i).data(Qt.UserRole) == new_template.template_id:
                    self.templates_list_widget.setCurrentRow(i);
                    break

    def delete_selected_template(self):
        current_item = self.templates_list_widget.currentItem()
        if not current_item: return
        template_id = current_item.data(Qt.UserRole)
        template_to_delete = next((t for t in self.templates_list_ref if t.template_id == template_id), None)
        if template_to_delete and QMessageBox.question(self, "Delete Template", f"Delete '{template_to_delete.name}'?",
                                                       QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.templates_list_ref.remove(template_to_delete)
            self.templates_updated.emit(self.templates_list_ref)
            self.load_and_display_templates_list()

    def on_template_selected(self):
        current_item = self.templates_list_widget.currentItem()
        if not current_item: self.clear_template_details_ui(); self.delete_template_btn.setEnabled(False); return
        template_id = current_item.data(Qt.UserRole)
        newly_selected_template = next((t for t in self.templates_list_ref if t.template_id == template_id), None)
        self.delete_template_btn.setEnabled(True)
        if newly_selected_template:
            if self.current_template and self.current_template.template_id != newly_selected_template.template_id:
                self.save_current_template_details(silent=True)
            self.current_template = newly_selected_template
            self.set_template_details_enabled(True)
            self.template_id_label.setText(self.current_template.template_id)
            self.template_name_edit.setText(self.current_template.name)
            self.keywords_edit.setText(", ".join(self.current_template.identifying_keywords))
            self.redactions_edit.setText(", ".join(self.current_template.redaction_texts))
            self.test_pdf_label.setText(os.path.basename(
                self.current_template.test_pdf_path) if self.current_template.test_pdf_path else "No Test PDF loaded")
            self.current_test_page_idx = 0
            if self.current_template.test_pdf_path and os.path.exists(self.current_template.test_pdf_path):
                self.template_pdf_viewer.load_pdf(self.current_template.test_pdf_path, self.current_test_page_idx,
                                                  self.current_template.redaction_texts)
            else:
                self.template_pdf_viewer.load_pdf(None)
            self.update_viewer_page_label()
            self.load_field_definitions_for_template()
        else:
            self.clear_template_details_ui()

    def clear_template_details_ui(self):  # Renamed
        self.current_template = None
        self.template_id_label.setText("N/A");
        self.template_name_edit.clear()
        self.keywords_edit.clear();
        self.test_pdf_label.setText("No Test PDF loaded")
        self.redactions_edit.clear();
        self.template_pdf_viewer.load_pdf(None)
        self.update_viewer_page_label();
        self.fields_table.setRowCount(0)
        self.set_template_details_enabled(False)
        self.edit_field_btn.setEnabled(False);
        self.delete_field_btn.setEnabled(False);
        self.test_all_fields_btn.setEnabled(False)

    def save_current_template_details(self, silent=False):
        if not self.current_template: return
        name_changed = self.current_template.name != self.template_name_edit.text()
        self.current_template.name = self.template_name_edit.text()
        self.current_template.identifying_keywords = [k.strip() for k in self.keywords_edit.text().split(',') if
                                                      k.strip()]
        new_redactions = [r.strip() for r in self.redactions_edit.text().split(',') if r.strip()]
        redactions_changed = (set(self.current_template.redaction_texts) != set(new_redactions))
        self.current_template.redaction_texts = new_redactions
        if not silent or name_changed or redactions_changed: self.templates_updated.emit(self.templates_list_ref)
        if name_changed:
            current_list_item = self.templates_list_widget.currentItem()
            if current_list_item and current_list_item.data(Qt.UserRole) == self.current_template.template_id:
                current_list_item.setText(f"{self.current_template.name} (ID: {self.current_template.template_id})")
        if redactions_changed and self.template_pdf_viewer.pdf_processor and self.template_pdf_viewer.pdf_processor.doc:
            self.template_pdf_viewer.update_redactions(self.current_template.redaction_texts)

    def load_test_pdf_for_template(self):
        if not self.current_template: QMessageBox.warning(self, "Load Test PDF", "Select template first."); return
        pdf_path, _ = QFileDialog.getOpenFileName(self, "Select Test PDF", "", "PDF Files (*.pdf)")
        if pdf_path:
            self.current_template.test_pdf_path = pdf_path
            self.test_pdf_label.setText(os.path.basename(pdf_path))
            self.current_test_page_idx = 0
            self.template_pdf_viewer.load_pdf(pdf_path, self.current_test_page_idx,
                                              self.current_template.redaction_texts)
            self.update_viewer_page_label()
            self.save_current_template_details(silent=True)  # Save path
            self.test_all_fields_btn.setEnabled(
                self.template_pdf_viewer.pdf_processor and self.template_pdf_viewer.pdf_processor.doc is not None)

    def viewer_prev_page(self):
        if self.template_pdf_viewer.pdf_processor and self.template_pdf_viewer.pdf_processor.doc and self.current_test_page_idx > 0:
            self.current_test_page_idx -= 1
            self.template_pdf_viewer.display_page(self.current_test_page_idx)
            self.update_viewer_page_label();
            self.template_pdf_viewer.clear_highlights()

    def viewer_next_page(self):
        if self.template_pdf_viewer.pdf_processor and self.template_pdf_viewer.pdf_processor.doc and self.current_test_page_idx < self.template_pdf_viewer.pdf_processor.get_page_count() - 1:
            self.current_test_page_idx += 1
            self.template_pdf_viewer.display_page(self.current_test_page_idx)
            self.update_viewer_page_label();
            self.template_pdf_viewer.clear_highlights()

    def update_viewer_page_label(self):
        if self.template_pdf_viewer.pdf_processor and self.template_pdf_viewer.pdf_processor.doc:
            self.page_indicator_label.setText(
                f"Page: {self.current_test_page_idx + 1}/{self.template_pdf_viewer.pdf_processor.get_page_count()}")
        else:
            self.page_indicator_label.setText("Page: -/- (No Test PDF)")

    def load_field_definitions_for_template(self):  # Same as previous
        self.fields_table.setRowCount(0)
        self.edit_field_btn.setEnabled(False);
        self.delete_field_btn.setEnabled(False)
        can_test_all = (self.current_template and self.current_template.field_definitions and
                        self.template_pdf_viewer.pdf_processor and self.template_pdf_viewer.pdf_processor.doc is not None)
        self.test_all_fields_btn.setEnabled(can_test_all)
        if not self.current_template: return
        for fd in self.current_template.field_definitions:
            row = self.fields_table.rowCount();
            self.fields_table.insertRow(row)
            self.fields_table.setItem(row, 0, QTableWidgetItem(fd.label))
            self.fields_table.setItem(row, 1, QTableWidgetItem(fd.sub_type or "N/A"))
            anchor_key_text = ""
            if fd.sub_type == "Anchor Word":
                anchor_key_text = fd.anchor_text or ""
            elif fd.sub_type == "Anchor Point" and fd.anchor_point:
                anchor_key_text = f"Pt({fd.anchor_point[0]:.0f},{fd.anchor_point[1]:.0f})"
            elif fd.sub_type == "Coordinate Box" and fd.coordinate_box:
                cb = fd.coordinate_box; anchor_key_text = f"Box"
            self.fields_table.setItem(row, 2, QTableWidgetItem(anchor_key_text))
            self.fields_table.setItem(row, 3, QTableWidgetItem(fd.extraction_scope or "N/A"))
            self.fields_table.setItem(row, 4, QTableWidgetItem(str(fd.n_words) if fd.n_words > 0 and (
                        "N Words" in (fd.extraction_scope or "") or "Value After Anchor (Next N Words)" in (
                            fd.extraction_scope or "")) else "-"))
            self.fields_table.item(row, 0).setData(Qt.UserRole, fd.field_id)

    def on_field_selection_changed(self):
        is_selected = bool(self.fields_table.selectedItems())
        self.edit_field_btn.setEnabled(is_selected);
        self.delete_field_btn.setEnabled(is_selected)

    def add_new_field_definition(self):
        if not self.current_template: QMessageBox.warning(self, "Add Field", "No template selected."); return
        if not self.template_pdf_viewer.pdf_processor or not self.template_pdf_viewer.pdf_processor.doc:
            QMessageBox.information(self, "Add Field", "Load a Test PDF for the template first.");
            return
        self.active_field_dialog = FieldExtractionDialog(self.current_template, None, self, self)
        self.active_field_dialog.fieldDefinitionSaved.connect(self.handle_field_saved)
        self.active_field_dialog.exec_()
        self.active_field_dialog = None  # Clear active dialog
        self.template_pdf_viewer.set_selection_mode(None)  # Reset main viewer mode after dialog closes

    def edit_selected_field_definition(self):
        if not self.current_template or not self.fields_table.selectedItems(): return
        if not self.template_pdf_viewer.pdf_processor or not self.template_pdf_viewer.pdf_processor.doc:
            QMessageBox.information(self, "Edit Field", "Load a Test PDF for the template first.");
            return
        field_id = self.fields_table.item(self.fields_table.currentRow(), 0).data(Qt.UserRole)
        field_to_edit = next((fd for fd in self.current_template.field_definitions if fd.field_id == field_id), None)
        if field_to_edit:
            self.active_field_dialog = FieldExtractionDialog(self.current_template, field_to_edit, self, self)
            self.active_field_dialog.fieldDefinitionSaved.connect(self.handle_field_saved)
            self.active_field_dialog.exec_()
            self.active_field_dialog = None
            self.template_pdf_viewer.set_selection_mode(None)

    def handle_field_saved(self, saved_field_def: FieldExtractionDefinition):  # Same as previous
        if not self.current_template: return
        existing_idx = next((i for i, fd in enumerate(self.current_template.field_definitions) if
                             fd.field_id == saved_field_def.field_id), -1)
        if existing_idx != -1:
            self.current_template.field_definitions[existing_idx] = saved_field_def
        else:
            self.current_template.field_definitions.append(saved_field_def)
        self.save_current_template_details(silent=True)
        self.templates_updated.emit(self.templates_list_ref)
        self.load_field_definitions_for_template()

    def delete_selected_field_definition(self):  # Same as previous
        if not self.current_template or not self.fields_table.selectedItems(): return
        field_id = self.fields_table.item(self.fields_table.currentRow(), 0).data(Qt.UserRole)
        field_to_delete = next((fd for fd in self.current_template.field_definitions if fd.field_id == field_id), None)
        if field_to_delete and QMessageBox.question(self, "Delete Field", f"Delete field '{field_to_delete.label}'?",
                                                    QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.current_template.field_definitions.remove(field_to_delete)
            self.save_current_template_details(silent=True)
            self.templates_updated.emit(self.templates_list_ref)
            self.load_field_definitions_for_template()

    def test_all_fields_on_current_page(self):  # Same as previous
        if not self.current_template or not self.template_pdf_viewer.pdf_processor or not self.template_pdf_viewer.pdf_processor.doc:
            QMessageBox.warning(self, "Test All Fields", "No template or Test PDF loaded.");
            return
        self.template_pdf_viewer.clear_highlights();
        results_text_list = [];
        any_errors = False
        processor_to_use = self.template_pdf_viewer.pdf_processor
        for field_def in self.current_template.field_definitions:
            try:
                result = processor_to_use.extract_field_data(field_def, self.current_test_page_idx)
                results_text_list.append(
                    f"<b>{field_def.label}:</b> {result.extracted_text if result.extracted_text else '<i>(empty)</i>'}")
                if "[ERROR" in result.extracted_text.upper(): any_errors = True
                self.template_pdf_viewer.add_highlight_multi(result.highlight_rects_pdf_coords,
                                                             result.anchor_highlight_rects_pdf_coords)
            except Exception as e:
                results_text_list.append(f"<b>{field_def.label}:</b> [TEST ERROR: {e}]"); any_errors = True
        msg_box = QMessageBox(self);
        msg_box.setWindowTitle("Test All Fields Results (Current Page)");
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setText("<br>".join(results_text_list))
        msg_box.setIcon(QMessageBox.Warning if any_errors else QMessageBox.Information)
        msg_box.setInformativeText(
            "Some fields had errors or no extraction." if any_errors else "Extraction test complete.")
        msg_box.exec_()

    # Callbacks for FieldExtractionDialog to interact with main viewer
    def set_main_viewer_selection_mode(self, mode, listener_dialog: FieldExtractionDialog):
        self.active_field_dialog = listener_dialog  # Store which dialog is listening
        self.template_pdf_viewer.set_selection_mode(mode)
        if mode: QMessageBox.information(self, f"Select {mode}",
                                         f"Click in the main PDF viewer to select a {mode} for the field '{listener_dialog.current_field_def.label or 'New Field'}'.")

    def on_main_viewer_point_selected(self, pdf_point: QPointF):
        if self.active_field_dialog:
            self.active_field_dialog.update_anchor_point_from_main_viewer(pdf_point)
            # Highlight the point on the main viewer
            self.template_pdf_viewer.clear_highlights()
            self.template_pdf_viewer.add_highlight_multi(None,
                                                         [(pdf_point.x() - 1, pdf_point.y() - 1, pdf_point.x() + 1,
                                                           pdf_point.y() + 1)], anchor_color=QColor(HSBC_RED),
                                                         opacity=0.9)
        self.template_pdf_viewer.set_selection_mode(None)  # Reset mode

    def on_main_viewer_box_selected(self, pdf_rect: QRectF):
        if self.active_field_dialog:
            self.active_field_dialog.update_coord_box_from_main_viewer(pdf_rect)
            self.template_pdf_viewer.clear_highlights()
            self.template_pdf_viewer.add_highlight_multi(
                [(pdf_rect.left(), pdf_rect.top(), pdf_rect.right(), pdf_rect.bottom())])
        self.template_pdf_viewer.set_selection_mode(None)

    def closeEvent(self, event):
        if self.pdf_processor: self.pdf_processor.close()
        super().closeEvent(event)


# ReportViewerDialog should remain largely the same as the last fully correct version.
# Ensure its __init__ takes DocumentRunReport.
# ... (ReportViewerDialog from previous fully correct response) ...
class ReportViewerDialog(QDialog):  # For DocumentRunReport
    reportUpdated = pyqtSignal(DocumentRunReport)

    def __init__(self, doc_run_report: DocumentRunReport, base_pdf_folder_path: str, parent=None):
        super().__init__(parent)
        self.doc_run_report = doc_run_report
        self.base_pdf_folder_path = base_pdf_folder_path if base_pdf_folder_path else ""
        self.pdf_processor_local = PDFProcessor()

        self.setWindowTitle(
            f"Report: {doc_run_report.source_pdf_path or 'N/A'} (Template: {doc_run_report.applied_template_id or 'N/A'})")
        self.setMinimumSize(1100, 750);
        self.setStyleSheet(f"QDialog {{ background-color: {LIGHT_GRAY_BLUE}; }}")
        self.main_layout = QHBoxLayout(self);
        self.splitter = QSplitter(Qt.Horizontal)
        pdf_viewer_panel = QWidget();
        pdf_viewer_layout = QVBoxLayout(pdf_viewer_panel)
        page_nav_layout = QHBoxLayout();
        self.prev_page_button = QToolButton();
        self.prev_page_button.setIcon(
            self.style().standardIcon(getattr(self.style(), ICON_PREV, self.style().SP_ArrowLeft)));
        self.prev_page_button.clicked.connect(self.prev_report_page)
        self.page_label = QLabel("Page: -/-");
        self.next_page_button = QToolButton();
        self.next_page_button.setIcon(
            self.style().standardIcon(getattr(self.style(), ICON_NEXT, self.style().SP_ArrowRight)));
        self.next_page_button.clicked.connect(self.next_report_page)
        page_nav_layout.addWidget(self.prev_page_button);
        page_nav_layout.addStretch();
        page_nav_layout.addWidget(self.page_label);
        page_nav_layout.addStretch();
        page_nav_layout.addWidget(self.next_page_button)
        pdf_viewer_layout.addLayout(page_nav_layout);
        self.pdf_viewer_widget = PDFViewerWidget();
        pdf_viewer_layout.addWidget(self.pdf_viewer_widget);
        self.splitter.addWidget(pdf_viewer_panel)
        self.results_widget = QWidget();
        self.results_layout = QVBoxLayout(self.results_widget)
        self.results_table = QTableWidget();
        self.results_table.setColumnCount(3);
        self.results_table.setHorizontalHeaderLabels(["Field Label", "Extracted Value", "Page"])
        self.results_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents);
        self.results_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch);
        self.results_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.results_table.setSelectionBehavior(QAbstractItemView.SelectRows);
        self.results_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.results_table.itemSelectionChanged.connect(self.on_report_result_selected);
        self.results_table.itemChanged.connect(self.on_report_result_edited)
        self.results_layout.addWidget(QLabel(f"Report ID: {doc_run_report.run_id} ({doc_run_report.timestamp})"));
        self.results_layout.addWidget(self.results_table)
        export_layout = QHBoxLayout();
        self.export_csv_button = QPushButton("Export to CSV");
        self.export_csv_button.clicked.connect(self.export_report_to_csv)
        self.export_excel_button = QPushButton("Export to Excel");
        self.export_excel_button.clicked.connect(self.export_report_to_excel)
        export_layout.addStretch();
        export_layout.addWidget(self.export_csv_button);
        export_layout.addWidget(self.export_excel_button);
        self.results_layout.addLayout(export_layout)
        self.splitter.addWidget(self.results_widget);
        self.main_layout.addWidget(self.splitter);
        self.splitter.setSizes([650, 450])
        self._data_changed_in_session = False;
        self._current_report_page_idx = 0
        self._load_pdf_for_report();
        self._populate_results_table()

    def _load_pdf_for_report(self):
        if not self.doc_run_report.source_pdf_path: QMessageBox.warning(self, "PDF Missing",
                                                                        "Source PDF info missing."); return
        full_pdf_path = os.path.join(self.base_pdf_folder_path, self.doc_run_report.source_pdf_path)
        self.pdf_processor_local.load_pdf(full_pdf_path)
        if self.pdf_processor_local.doc:
            self.pdf_viewer_widget.pdf_processor = self.pdf_processor_local
            self._current_report_page_idx = self.doc_run_report.field_results[
                0].source_page_idx if self.doc_run_report.field_results else 0
            self.pdf_viewer_widget.display_page(self._current_report_page_idx)
        else:
            QMessageBox.warning(self, "PDF Load Error", f"Could not load PDF: {full_pdf_path}")
        self.update_report_page_label()

    def _populate_results_table(self):
        self.results_table.setRowCount(0);
        self.results_table.blockSignals(True)
        for row, res_item in enumerate(self.doc_run_report.field_results):
            self.results_table.insertRow(row)
            lbl_item = QTableWidgetItem(res_item.field_label);
            lbl_item.setFlags(lbl_item.flags() & ~Qt.ItemIsEditable);
            self.results_table.setItem(row, 0, lbl_item)
            val_item = QTableWidgetItem(str(res_item.extracted_text));
            val_item.setData(Qt.UserRole, res_item);
            self.results_table.setItem(row, 1, val_item)
            pg_item = QTableWidgetItem(str(res_item.source_page_idx + 1));
            pg_item.setFlags(pg_item.flags() & ~Qt.ItemIsEditable);
            self.results_table.setItem(row, 2, pg_item)
        self.results_table.blockSignals(False)
        if self.results_table.rowCount() > 0: self.results_table.selectRow(0)

    def on_report_result_selected(self):
        sel_rows = self.results_table.selectionModel().selectedRows()
        if not sel_rows: self.pdf_viewer_widget.clear_highlights(); return
        res_item: ExtractedFieldResult = self.results_table.item(sel_rows[0].row(), 1).data(Qt.UserRole)
        if res_item and self.pdf_viewer_widget.pdf_processor and self.pdf_viewer_widget.pdf_processor.doc and self.pdf_viewer_widget.current_page_idx != res_item.source_page_idx:
            self._current_report_page_idx = res_item.source_page_idx
            self.pdf_viewer_widget.display_page(self._current_report_page_idx);
            self.update_report_page_label()
        self.pdf_viewer_widget.clear_highlights()
        if res_item: self.pdf_viewer_widget.add_highlight_multi(res_item.highlight_rects_pdf_coords,
                                                                res_item.anchor_highlight_rects_pdf_coords)
        if res_item and res_item.highlight_rects_pdf_coords:
            scene_r = self.pdf_viewer_widget.pdf_to_scene_coords_rect(res_item.highlight_rects_pdf_coords[0])
            if scene_r: self.pdf_viewer_widget.ensureVisible(scene_r, 50, 50)

    def on_report_result_edited(self, item: QTableWidgetItem):
        if item.column() == 1:
            res_item: ExtractedFieldResult = item.data(Qt.UserRole)
            if res_item and res_item.extracted_text != item.text():
                res_item.extracted_text = item.text();
                self._data_changed_in_session = True;
                item.setBackground(QColor(MEDIUM_BLUE).lighter(180))

    def prev_report_page(self):
        if self.pdf_processor_local and self.pdf_processor_local.doc and self._current_report_page_idx > 0:
            self._current_report_page_idx -= 1;
            self.pdf_viewer_widget.display_page(self._current_report_page_idx);
            self.update_report_page_label()

    def next_report_page(self):
        if self.pdf_processor_local and self.pdf_processor_local.doc and self._current_report_page_idx < self.pdf_processor_local.get_page_count() - 1:
            self._current_report_page_idx += 1;
            self.pdf_viewer_widget.display_page(self._current_report_page_idx);
            self.update_report_page_label()

    def update_report_page_label(self):
        self.page_label.setText(
            f"Page: {self._current_report_page_idx + 1}/{self.pdf_processor_local.get_page_count()}" if self.pdf_processor_local and self.pdf_processor_local.doc else "Page: -/- (No PDF)")

    def accept(self):
        if self._data_changed_in_session: self.reportUpdated.emit(self.doc_run_report)
        super().accept()

    def reject(self):
        if self._data_changed_in_session and QMessageBox.question(self, "Unsaved Changes", "Discard changes?",
                                                                  QMessageBox.Yes | QMessageBox.No) == QMessageBox.No: return
        super().reject()

    def closeEvent(self, event):
        if self.pdf_processor_local: self.pdf_processor_local.close()
        super().closeEvent(event)

    def _get_all_results_for_export(self):
        return [(res.field_label, res.extracted_text) for res in self.doc_run_report.field_results]

    def export_report_to_csv(self):  # same as previous
        data_to_export = self._get_all_results_for_export()
        if not data_to_export: QMessageBox.information(self, "Export CSV", "No data."); return
        default_filename = f"report_{os.path.splitext(self.doc_run_report.source_pdf_path or 'file')[0]}_{self.doc_run_report.run_id}.csv"
        path, _ = QFileDialog.getSaveFileName(self, "Save CSV", default_filename, "CSV Files (*.csv)")
        if path:
            try:
                import csv
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f);
                    writer.writerow(["Field Name", "Extracted Value"])
                    writer.writerows(data_to_export)
                QMessageBox.information(self, "Export CSV", f"Exported to {path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"CSV Export Failed: {e}")

    def export_report_to_excel(self):  # same as previous
        data_to_export = self._get_all_results_for_export()
        if not data_to_export: QMessageBox.information(self, "Export Excel", "No data."); return
        default_filename = f"report_{os.path.splitext(self.doc_run_report.source_pdf_path or 'file')[0]}_{self.doc_run_report.run_id}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", default_filename, "Excel Files (*.xlsx)")
        if path:
            try:
                from openpyxl import Workbook;
                from openpyxl.styles import Font, PatternFill;
                from openpyxl.utils import get_column_letter
                wb = Workbook();
                ws = wb.active;
                ws.title = "Extraction Report";
                start_row, start_col = 3, 3
                hdr_font = Font(bold=True);
                fn_fill = PatternFill(start_color="D9D9D9", fill_type="solid")
                ws.cell(row=start_row, column=start_col, value="Field Name").font = hdr_font;
                ws.cell(row=start_row, column=start_col).fill = fn_fill
                ws.cell(row=start_row, column=start_col + 1, value="Extracted Value").font = hdr_font
                for r_idx, (label, value) in enumerate(data_to_export):
                    cr = start_row + 1 + r_idx
                    c_fn = ws.cell(row=cr, column=start_col, value=label);
                    c_fn.fill = fn_fill;
                    c_fn.font = Font(bold=True)
                    ws.cell(row=cr, column=start_col + 1, value=value)
                for col_idx_offset, hdr_txt in enumerate(["Field Name", "Extracted Value"]):
                    col_char = get_column_letter(start_col + col_idx_offset);
                    max_len = len(hdr_txt)
                    for r_data in data_to_export: max_len = max(max_len, len(str(r_data[col_idx_offset])))
                    ws.column_dimensions[col_char].width = max_len + 5
                wb.save(path);
                QMessageBox.information(self, "Export Excel", f"Exported to {path}")
            except ImportError:
                QMessageBox.critical(self, "Error", "openpyxl needed.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Excel Export Failed: {e}")